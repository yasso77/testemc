from collections import defaultdict
from datetime import datetime
from urllib.parse import urlencode
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.core.paginator import Paginator
from django.db.models import (
    Count, Q, Max, Subquery, OuterRef,Exists,
    Case, Value, When, DateField, IntegerField
)
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Prefetch
from manager import models
from manager.decorators import permission_required_with_redirect
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from django.template.loader import render_to_string
from manager.model.patient import AgentCompany, CallTrack, City, Patient, PatientMedicalHistory
from manager.model.visit import PatientVisits
from manager.orm import ORMPatientsHandling
from django.views.generic.list import ListView
from django.utils import timezone
from django.http import HttpResponse
from django.utils.timezone import localtime
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.db.models import Count
from django.shortcuts import render
from django.db.models import Count, Q
from django.contrib.auth.models import User
from django.db.models import Prefetch
from django.db.models import Sum

context = {
    'users': User.objects.filter(
        is_active=True,
        groups__name__iexact="call center"
    ).distinct()
}


ormObj=ORMPatientsHandling()

class ReportView(ListView):

  

    @permission_required_with_redirect('manager.LiveReport', login_url='/no-permission/')
    def LiveDegreeReport(request):
        current_time = datetime.now().time()
        return render(request,'LiveDegreeReport.html',context={'current_time': current_time})

    @permission_required_with_redirect('manager.LiveReport', login_url='/no-permission/')
    
    def ajaxReportChartEvlDegree(request):
        if request.method == 'POST' and 'inputDate' in request.POST:
                

                today = timezone.now().date() 
                
                visitdate = request.POST.get('inputDate', None)
                queryset = PatientVisits.objects.filter(visitdate=today).values('evaluationeegree')  

            # Convert QuerySet to list of dictionaries
                data = list(queryset.values())
               

            # Return JsonResponse with the converted data
                return JsonResponse(data, safe=False)
                
        else:
                return JsonResponse({'success': False, 'error': 'Invalid request'})  
            
    @login_required
    @permission_required_with_redirect('manager.ShowPatientReport',login_url='/no-permission/')
    def showPatientData(request):
            
        patientList=ormObj.getPatientsTodayWithNoVisitYet()
        patientcount=patientList.count()  
        
        return render(request, 'SearchOnPatients.html', {
                'patients': patientList,
                'Total': patientcount
                })

    @permission_required_with_redirect('manager.UpdatePatinetData', login_url='/no-permission/')
    def uploadedPatientDataList(request): 
            
        if request.method == 'POST':
                fromDate = request.POST.get('txtFromDate')
                toDate = request.POST.get('txtToDate')
                # Fetch patients with their related visits
                queryset = ormObj.get_all_Patients_Between_Period(fromDate, toDate)
                context = {
                    'patients': queryset,
                }
        else:
                context = {}  # Define an empty context if request method is not POST

        return render(request, 'UploadedPatientsList.html', context)
    
    

    def compare_visits(request):
        # Fetch all relevant visits and include patient info
        visits = PatientVisits.objects.select_related('patientid', 'classifiedID').filter(
            visittype__in=['A', 'D']
        )

        # Organize visits by patient and visit type
        patient_data = {}
        for visit in visits:
            pid = visit.patientid
            if pid not in patient_data:
                patient_data[pid] = {
                    'name': visit.patientid.fullname,
                    'fileserial': visit.patientid.fileserial,
                    'A': None,
                    'D': None,
                }
            patient_data[pid][visit.visittype] = {
                'evaluation': visit.evaluationeegree,
                'classification': visit.classifiedID.classifiedCategory if visit.classifiedID else 'N/A'
            }

        # Build comparison list
        comparison_data = []
        for pid, data in patient_data.items():
            if data['A'] and data['D']:
                comparison_data.append({
                    'patient_id': pid,
                    'patient_name': data['name'],
                    'fileserial': data['fileserial'],
                    'a_evaluation': data['A']['evaluation'],
                    'd_evaluation': data['D']['evaluation'],
                    'a_classification': data['A']['classification'],
                    'd_classification': data['D']['classification'],
                    'match': '✅ Match' if data['A']['evaluation'] == data['D']['evaluation'] and
                                            data['A']['classification'] == data['D']['classification']
                            else '❌ No Match'
                })

        return render(request, 'reports/compare_visits.html', {'comparison_data': comparison_data})

    
 
    def patient_report_view(request):
        # =========================
        # Read filters from request
        # =========================
        date_field = request.GET.get('date_field', 'createdDate')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        city_id = request.GET.get('city')
        agent_id = request.GET.get('agentID')
        lead_source = request.GET.get('leadSource')
        user_id = request.GET.get('users')
        export = request.GET.get('export')

        # =========================
        # Detect if any filter is applied
        # =========================
        has_filters = any([
            date_from,
            date_to,
            city_id,
            agent_id,
            lead_source,
            user_id,
        ])

        # =========================
        # Start with EMPTY queryset
        # =========================
        patients = Patient.objects.none()

        # =========================
        # Apply filters only if user filtered
        # =========================
        if has_filters:
            patients = Patient.objects.filter(isDeleted=False)

            # ---- Date filtering
            if date_from and date_to:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')

                    if date_field == 'createdDate':
                        patients = patients.filter(
                            createdDate__range=(
                                date_from_obj,
                                date_to_obj + timedelta(days=1)
                            )
                        )
                    else:
                        patients = patients.filter(
                            **{
                                f"{date_field}__range": (
                                    date_from_obj.date(),
                                    date_to_obj.date()
                                )
                            }
                        )
                except ValueError:
                    pass

            # ---- Other filters
            if city_id:
                patients = patients.filter(city_id=city_id)

            if agent_id:
                patients = patients.filter(agentID_id=agent_id)

            if lead_source:
                patients = patients.filter(leadSource=lead_source)

            if user_id:
                patients = patients.filter(createdBy_id=user_id)

            # ---- Optimize FK loading
            patients = patients.select_related(
                'city',
                'agentID',
                'createdBy'
            )

        # =========================
        # Excel Export (only if filtered)
        # =========================
        if export == 'excel' and has_filters:
            wb = Workbook()
            ws = wb.active
            ws.title = "Patients Report"

            headers = [
                'Code', 'File Serial', 'Name',                 'Lead Source', 'Agent', 'Suffered Case',
                'City', 'Created By', 'Created Date', 'Attendance Date'
            ]
            ws.append(headers)

            for p in patients:
                ws.append([
                    p.reservationCode,
                    p.fileserial,
                    p.fullname,
                  
                    p.leadSource,
                    p.agentID.AgentCompany if p.agentID else '',
                    str(p.sufferedcase),
                    p.city.cityName if p.city else '',
                    str(p.createdBy),
                    localtime(p.createdDate).strftime('%Y-%m-%d %H:%M') if p.createdDate else '',
                    p.attendanceDate.strftime('%Y-%m-%d') if p.attendanceDate else '',
                ])

            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=patients_report.xlsx'
            wb.save(response)
            return response

        # =========================
        # Pagination (safe & fast)
        # =========================
        paginator = Paginator(patients.order_by('-createdDate'), 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # =========================
        # Static choices
        # =========================
        leadSource_Choices = [
            ('Facebook', 'Facebook'),
            ('Whatsapp', 'Whatsapp'),
            ('Youtube', 'Youtube'),
            ('Newspaper', 'Newspaper'),
            ('Friend', 'Friend'),
            ('Call', 'Call'),
            ('Instagram', 'Instagram'),
            ('Center', 'Center'),
        ]

        # =========================
        # Preserve query params
        # =========================
        get_params = request.GET.copy()
        get_params.pop('page', None)
        query_string = get_params.urlencode()

        # =========================
        # Context
        # =========================
        context = {
            'page_obj': page_obj,
            'total_count': paginator.count if has_filters else 0,
            'has_filters': has_filters,

            'cities': City.objects.all(),
            'agents': AgentCompany.objects.all(),
            'users': User.objects.filter(
                Q(groups__name__iexact="Call Center") |
                Q(groups__name__iexact="Reception")
            ).distinct().order_by('username'),

            'date_field': date_field,
            'date_from': date_from,
            'date_to': date_to,
            'city_id': str(city_id) if city_id else '',
            'agent_id': str(agent_id) if agent_id else '',
            'user_id': str(user_id) if user_id else '',
            'lead_source': lead_source or '',
            'lead_sources': dict(leadSource_Choices),

            'query_string': query_string,
        }

        return render(request, 'reports/export_patients_xl.html', context)

    
   
    
    def showPatientDataAttendedToday(request):
        
        today = timezone.localdate()   # ✅ DEFINE today FIRST

         # Default to today if not provided
        from_date = request.GET.get('from_date', today)
        to_date = request.GET.get('to_date', today)

        # 🚫 Hard stop if dates are missing
        if not from_date or not to_date:
            return render(request, 'reports/whoAttend.html', {
                'patients': [],
                'Total': 0,
                'error': 'Please select FROM and TO dates'
            })

        visit_qs = PatientVisits.objects.filter(
            visitdate__gte=from_date,
            visitdate__lte=to_date
        )

        patientList = (
            Patient.objects
            .filter(
                attendanceDate__gte=from_date,
                attendanceDate__lte=to_date,
                formPrinted=1
            )
            .prefetch_related(
                Prefetch('patientvisits', queryset=visit_qs)
            )
            .order_by('-attendanceDate')
        )

        return render(request, 'reports/whoAttend.html', {
            'patients': patientList,
            'Total': patientList.count(),
            'from_date': from_date,
            'to_date': to_date
        })

    
   


    def visit_evaluation_pivot(request):
        today = timezone.localdate()

        from_date = request.GET.get('from_date', today)
        to_date = request.GET.get('to_date', today)

        data = (
            PatientVisits.objects
            .filter(
                visittype='A',
                VisitDate__date__gte=from_date,
                VisitDate__date__lte=to_date
            )
            .values('VisitDate__date')
            .annotate(
                eval_plus=Count(
                    Case(When(EvaluationDegree='++', then=1), output_field=IntegerField())
                ),
                eval_surgery=Count(
                    Case(When(EvaluationDegree='Surgery', then=1), output_field=IntegerField())
                ),
                eval_6=Count(
                    Case(When(EvaluationDegree='6/6', then=1), output_field=IntegerField())
                ),
                eval_ok=Count(
                    Case(When(EvaluationDegree='OK', then=1), output_field=IntegerField())
                ),
                eval_bad=Count(
                    Case(When(EvaluationDegree='Bad', then=1), output_field=IntegerField())
                ),
                total=Count('id')
            )
            .order_by('VisitDate__date')
        )

        return render(request, 'reports/auditsStats.html', {
            'data': data,
            'from_date': from_date,
            'to_date': to_date
        })


        
        
    def doctors_stats(request):
        today = timezone.localdate()   # ✅ DEFINE today FIRST
        
        doctor_id = request.GET.get("doctor")  # <-- new filter
       # Base queryset
       
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        from_date = (
            datetime.strptime(from_date_str, '%Y-%m-%d').date()
            if from_date_str else today
        )
        to_date = (
            datetime.strptime(to_date_str, '%Y-%m-%d').date()
            if to_date_str else today
        )
        visits = PatientVisits.objects.filter(visittype="D")

        # Apply date filter
        if from_date and to_date:
            visits = visits.filter(visitdate__range=[from_date, to_date])

        # Apply doctor filter
        if doctor_id:
            visits = visits.filter(doctorid=doctor_id)

        # Aggregate stats
        stats = visits.values("doctorid", "doctorid__username").annotate(
            total_visits=Count("visitid"),
            surgery=Count("visitid", filter=Q(evaluationeegree="Surgery")),
            patients_lose=Count("visitid", filter=Q(evaluationeegree="Bad")),
        )

        # Build comparison data
        comparison_data = []
        for row in stats:
            surgery = row["surgery"]
            bad = row["patients_lose"]
            total = row["total_visits"]

            if total > 0:
                surgery_percent = (surgery / total) * 100
                bad_percent = (bad / total) * 100
            else:
                surgery_percent = bad_percent = 0

            if surgery_percent >= 70:
                status = "Good"
            elif surgery_percent >= 40:
                status = "Average"
            else:
                status = "Bad"

            comparison_data.append({
                "doctor_id": row["doctorid"],
                "doctor_name": row["doctorid__username"],
                "total_visits": total,
                "surgery": surgery,
                "patients_lose": bad,
                "stats": status,
                "surgery_percent": surgery_percent,
                "bad_percent": bad_percent,
            })

        # Get all doctors for dropdown
        doctors = User.objects.filter(groups__name="Doctors") if hasattr(User, 'groups') else User.objects.all()

        return render(request, "reports/doctorsStats.html", {
            "comparison_data": comparison_data,
            "from_date": from_date,
            "to_date": to_date,
            "doctors": doctors,
    })
        
    def patients_by_type(request, doctor_id, patient_type, from_date, to_date):
        
        if from_date == "null" or to_date == "null":
            return render(request, "reports/_patientsTable.html", {
                "visits": [],
                "title": "No Date Range Specified",
            })
        else:
            visits = PatientVisits.objects.filter(
                doctorid=doctor_id,
                visittype="D",
                visitdate__range=[from_date, to_date]
            )

            if patient_type == "surgery":
                visits = visits.filter(evaluationeegree="Surgery")
                title = "Surgery Patients"
            elif patient_type == "loss":
                visits = visits.filter(evaluationeegree="Bad")
                title = "Lost Patients"
            else:
                visits = []
                title = "Invalid Type"

            # Return only the patient table, not the full layout
            return render(request, "reports/_patientsTable.html", {
                "visits": visits,
                "title": title,
                "from_date": from_date,
                "to_date": to_date,
        })
    
    @require_POST       
    def delete_patient(request, patientid):
        patient = get_object_or_404(Patient, patientid=patientid)
        patient.updatedby = request.user.id
        patient.latestupdate = timezone.now()
        patient.isDeleted = True
        patient.save()

        search_text = request.POST.get("strText", "")

        base_url = reverse("SearchOnPatientResult")
        return redirect(f"{base_url}?{urlencode({'strText': search_text})}")
    
   #@require_GET      
    def delete_patientDuplicated(request, patientid,date_from,date_to):
        patient = get_object_or_404(Patient, patientid=patientid)
        patient.updatedby = request.user
        patient.latestupdate = timezone.now()
        patient.isDeleted = True
        patient.save()

        

        base_url = reverse("duplicatedPatients") + f"?date_from={date_from}&date_to={date_to}"
        
        return redirect(f"{base_url}")
     
        

            
                  
    # def delete_duplicatedpatient(request, patientid):
    #     patient = get_object_or_404(Patient, patientid=patientid)
    #     date_from_str = request.GET.get('date_from')
    #     date_to_str = request.GET.get('date_to')
        
    #     def parse_date(date_str):
    #         if not date_str:
    #             return today
    #         try:
    #             return datetime.strptime(date_str, '%Y-%m-%d').date()
    #         except ValueError:
    #             return today

    #     date_from = parse_date(date_from_str)
    #     date_to = parse_date(date_to_str)
        
    #     if request.method == 'GET':
    #         # Perform a soft delete by setting isDeleted to True
    #         patient.isDeleted = True
    #         patient.save()
    #         print("Deleted patient ID:", patientid)  # Debugging line
    #         return HttpResponseRedirect(    reverse('duplicatedPatients') )
    
     
    def reactivate_patient(request, patientid):
        patient = get_object_or_404(Patient, patientid=patientid)
        if request.method == 'GET':
            # Perform a soft delete by setting isDeleted to True
            patient.isDeleted = False
            patient.updatedby = request.user.id
            patient.latestupdate = timezone.now()
            patient.save()
           
            return redirect('archivedPatient')

    def confirm_page(request, patientName, fileserial):
        return render(
            request,
            "reports/ConfirmMsg.html",
            {
                "message": "The Reservation is updated Successfully.",
                "patientName": patientName,
                "fileserial": fileserial,
                #"patientName": patientName,
                #"show_print": True,
            },
        )

    def SearchOnPatient(request):
        return render(request, 'reports/SearchPatient.html') 
    
    def SearchOnPatientResult(request):
         
        strText = request.GET.get('strText', '')
       # print("Search Text:", strText)  # Debugging line to check the input value       
        recent_patients = (
            Patient.objects.active()
            .filter(
                Q(reservationCode__icontains=strText) |
                Q(mobile__icontains=strText) |  # Search in mobile
                Q(fileserial__icontains=strText) |  # Search in file serial
                Q(fullname__icontains=strText) |  # Search in name               
                Q(attendanceDate__icontains=strText)|  # Search in attendance date
                Q(createdDate__icontains=strText),
                #reservedBy=request.user  # Keep the reservedBy filter
            )
            .select_related('sufferedcase')
            .annotate(
               
                call_count=Count('call_patients'),  # Count number of call tracks for each patient
                last_call_date=Max('call_patients__createdDate'),  # Get the latest call date
                last_call_outcome=Subquery(
                    CallTrack.objects.filter(
                        patientID=OuterRef('pk')  # Reference the current patient
                    )
                    .values('outcome')[:1]  # Get the outcome of the latest call
                ),
                 has_medical_history=Exists(
                 PatientMedicalHistory.objects.filter(patient=OuterRef('pk'))
            )  
            )
            .values(
                'patientid','fileserial', 'fullname', 'reservationCode', 'leadSource',
                'createdDate','createdBy__username', 'city', 'mobile', 'age', 'sufferedcase__caseName',
                'sufferedcaseByPatient__caseName', 'expectedDate', 'gender', 'attendanceDate', 'birthdate',
                'call_count', 'last_call_date', 'last_call_outcome','has_medical_history'  # Add annotated fields
            )
        )  
        
        return render(request, 'reports/SearchResult.html', {'patients': recent_patients,'viewScope':strText,'total_count':recent_patients.count()})
    
    @login_required  
    def archivedPatientList(request):    

        # =========================
        # Filters
        # =========================      
        
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')       
        txtSearch = request.GET.get('txtSearch')     
       
        
        # =========================
        # Base queryset
        # =========================
        patients_qs = (
            Patient.objects.filter(isDeleted=True)
        )

        # =========================
        # Date filtering
        # =========================
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')

                patients_qs = patients_qs.filter(
                        createdDate__range=(
                            date_from_obj,
                            date_to_obj + timedelta(days=1)
                        )
                    )
              
            except ValueError:
                pass

        # =========================
        # Other filters
        # =========================
        if txtSearch:
            patients_qs = patients_qs.filter(
                Q(fullname__icontains=txtSearch) |
                Q(mobile__icontains=txtSearch)
            )        

        # =========================
        # Annotations & optimization
        # =========================
        patients_qs = patients_qs.select_related(
            'sufferedcase', 'createdBy', 'city', 'agentID'
        ).annotate(
            call_count=Count(
                'call_patients',
                filter=Q(call_patients__trackType='CC')
            ),
            last_call_date=Max(
                'call_patients__createdDate',
                filter=Q(call_patients__trackType='CC')
            ),
            last_call_outcome=Subquery(
                CallTrack.objects.filter(
                    patientID=OuterRef('pk'),
                    trackType='CC'
                )
                .order_by('-createdDate')
                .values('outcome')[:1]
            )
        ).order_by('-createdDate')

        # =========================
        # Pagination
        # =========================
        paginator = Paginator(patients_qs, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # =========================
        # AJAX response
        # =========================
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string(
                'reports/patient_rows_archived.html',
                {'patients': page_obj}
            )
            return JsonResponse({
                'html': html,
                'has_next': page_obj.has_next(),
                'total_count': paginator.count
            })

        # =========================
        # Normal render
         # =========================
      

        # =========================
        
      

        context = {
            'patients': page_obj,
            'total_count': paginator.count,  
            'date_from': date_from,
            'date_to': date_to,
           
        }
        return render(request, 'reports/archivedPatients.html', context)
    
    
    def audit_stats(request):
        today = timezone.localdate()

        # ✅ Always convert GET values to date objects
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        from_date = (
            datetime.strptime(from_date_str, '%Y-%m-%d').date()
            if from_date_str else today
        )
        to_date = (
            datetime.strptime(to_date_str, '%Y-%m-%d').date()
            if to_date_str else today
        )

        data = (
            PatientVisits.objects
            .filter(
                visittype='A',
                visitdate__gte=from_date,
                visitdate__lte=to_date
            )
            .values('visitdate')
            .annotate(
                eval_ok=Count(
                    Case(When(evaluationeegree='OK', then=1), output_field=IntegerField())
                ),
                eval_plus=Count(
                    Case(When(evaluationeegree='++', then=1), output_field=IntegerField())
                ),
                eval_surgery=Count(
                    Case(When(evaluationeegree='Surgery', then=1), output_field=IntegerField())
                ),
                eval_6=Count(
                    Case(When(evaluationeegree='6/6', then=1), output_field=IntegerField())
                ),
                eval_bad=Count(
                    Case(When(evaluationeegree='Bad', then=1), output_field=IntegerField())
                ),
                total=Count(
                Case(
                    When(evaluationeegree__isnull=False, then=1),
                    output_field=IntegerField()
                )
            )
        
            )
            .order_by('visitdate')
        )
        
        totals = data.aggregate(
        total_plus=Sum('eval_plus'),
        total_6_6=Sum('eval_6'),
        total_ok=Sum('eval_ok'),
        total_bad=Sum('eval_bad'),
        total_surgery=Sum('eval_surgery'),
        grand_total=Sum('total')
)
        
        chart_labels = ['++', '6/6', 'OK', 'Bad', 'Surgery']
        chart_values = [
            totals['total_plus'] or 0,
            totals['total_6_6'] or 0,
            totals['total_ok'] or 0,
            totals['total_bad'] or 0,
            totals['total_surgery'] or 0,
        ]



        return render(request, 'reports/auditStats.html', {
            'data': data,
            'totals': totals,
            'from_date': from_date,
            'to_date': to_date,
            'chart_labels': chart_labels,
            'chart_values': chart_values,
        })
        
        

    def duplicatedPatientList(request):

        # =========================
        # Read & parse dates safely
        # =========================
        today = timezone.localdate()

        def parse_date(date_str):
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                return today

        date_from = parse_date(request.GET.get('date_from'))
        date_to = parse_date(request.GET.get('date_to'))

        # =========================
        # Find duplicate keys
        # =========================
        duplicate_keys = (
            Patient.objects
            .filter(
                isDeleted=False,
                createdDate__range=(date_from, date_to)
            )
            .values('fullname', 'mobile')
            .annotate(cnt=Count('patientid'))
            .filter(cnt__gt=1)
        )

        # =========================
        # Get full duplicate records
        # =========================
        patients_qs = (
            Patient.objects
            .filter(
                isDeleted=False,
                attendanceDate__isnull=False,
                createdDate__range=(date_from, date_to)
            )
            .filter(
                Q(
                    fullname__in=duplicate_keys.values('fullname'),
                    mobile__in=duplicate_keys.values('mobile')
                )
            )
            .select_related('createdBy')
            .order_by('fullname', 'mobile', '-attendanceDate')
        )

        # =========================
        # Response
        # =========================
        context = {
            'patients': patients_qs,
            'total_count': patients_qs.count(),
            'date_from': date_from,
            'date_to': date_to,
        }

        return render(request, 'reports/duplicatedPatients.html', context)

        